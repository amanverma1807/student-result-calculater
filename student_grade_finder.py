# ============================================
# Student Grade Finder
# Use: to calculate result of students from excel file as input
# ============================================

# import of necessery libraries


from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import fonts, PatternFill
import os

# loading excel file as input


wb = load_workbook('C:\\python projects\\student grade finder.v1.1\\students number.xlsx')
ws = wb.active
# Adding new column headers


ws["E1"] = "Total"
ws["F1"] = "Percentage"
ws["G1"] = "Grade"
ws["H1"] = "Status"

# Calculating total and percentage for each student


for i in range(2,ws.max_row + 1):

    # Reading subject marks from columns B, C, D


    physics = ws['B'+str(i)].value
    maths = ws['C'+str(i)].value
    chem = ws['D'+str(i)].value

    # Calculating total and percentage


    total = physics + maths + chem
    ws['E'+str(i)] = total
    percentage = round((total/300)*100,2)
    ws["F"+str(i)] = percentage

# Assigning grades based on total marks


for i in range(2,ws.max_row + 1):
    num = ws['E'+str(i)].value
    if num >= 270:
        ws['G'+str(i)] = "A"
    elif num >= 240:
        ws['G'+str(i)] = "B"
    elif num >= 210:
        ws['G'+str(i)] = "C"
    elif num >= 180:
        ws['G'+str(i)] = "D"
    else:
        ws['G'+str(i)] = "F"

# Marking pass or fail with font color


for i in range(2,ws.max_row + 1):
    num = ws['E'+str(i)].value
    if num >= 180:
        ws['H'+str(i)] = "Pass"
        ws['H'+str(i)].font = fonts.Font(color='008000')
    else:
        ws['H'+str(i)] = "Fail"
        ws['H'+str(i)].font = fonts.Font(color='FF0000')

# Formatting header row - bold and colored


for i in range(1,ws.max_row + 2):
    char = get_column_letter(i)
    ws[char + '1'].font = fonts.Font(bold='true',color="384AB5")
    ws.column_dimensions[char].width = 20

#  Finding the topper


highest_total = 0 
topper_name = ''
for i in range(2,ws.max_row+1):
    total = ws['E'+str(i)].value
    if total > highest_total:
        highest_total = total
        topper_name = ws['A'+str(i)].value

        ws['J1'] = 'Topper name'
        ws['K1'] = 'Total'
        ws['J2'] = topper_name
        ws['K2'] = highest_total

# Highlighting fail students rows in red
#       

for row in range(2,ws.max_row+1):
    if ws['G'+str(row)].value == "F":
        for i in range(1,ws.max_column+1):
            char = get_column_letter(i)
            ws[char + str(row)].fill = PatternFill(start_color="FFC7CE",end_color="FFC7CE",fill_type='solid')

# saving final report


os.makedirs('output_folder', exist_ok=True)
wb.save('output_folder/final_report.xlsx')
